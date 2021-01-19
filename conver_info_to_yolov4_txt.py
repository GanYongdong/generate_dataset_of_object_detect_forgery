##从保存的所有信息的xlsx文件中，读取需要的并转换成对应yolo4的txtlabel文件

from openpyxl import load_workbook
import os
from pathlib import Path

#xlsx所在目录
path = '/home/ganyd/DataSet/dataset_irr_png/xlsx_all_info_path';

#获取目录下所有目录与文件名称
fileNameList = os.listdir(path)

#打开要写的txt文件，已有就清空，没有就创建
f = open('./yolov4Label.txt','w')
f.close()

for i in range(len(fileNameList)):

    xlsxFilePath = path + '/' + fileNameList[i]
    filePathTmp = Path(xlsxFilePath)
    if filePathTmp.is_dir():
        continue #如果这个是目录，不是文件则跳过

    #转换成jpg目录
    #picFilePath = fileNameList[i]
    picFilePath = fileNameList[i].split('.')[0]
    #picFilePath.replace('.xlsx', '.jpg')
    #picFilePath = '/home/ganyd/DataSet/dataset_irr_png/rgbOut/' + picFilePath + '.png'
    picFilePath = picFilePath + '.png'

    #第一列写图片所在目录
    strWillWrite = []
    if i != 0:
        strWillWrite = '\n' + picFilePath
    else:
        strWillWrite = picFilePath
    #print(picFilePath)

    #打开xlsx文件
    workbook = load_workbook(xlsxFilePath) #加载xlsx文件
    booksheet = workbook.active #获取活动的页
    #worksheet = workbook.get_sheet_by_name('Sheet1') #或者指定name获取sheet页
    rows = booksheet #获取sheet页的行数据
    columns = booksheet.columns #获取sheet页的列数据

    #循环获取每个目标信息
    #[left,top,w,h,cx,cy,right,bottom,area, cx_rate, cy_rate, w_rate, h_rate]
    j = 0
    for row in rows:
        j = j + 1
        xmin = booksheet.cell(row=j, column=4).value
        ymin = booksheet.cell(row=j, column=5).value
        xmax = booksheet.cell(row=j, column=10).value
        ymax = booksheet.cell(row=j, column=11).value
        classId = booksheet.cell(row=j, column=1).value
        strWillWrite = strWillWrite + ' ' + str(xmin) + ',' + str(ymin) + ',' + str(xmax) + ',' + str(ymax) + ',' + str(classId)
        #f.write(xmin,xmin,xmin,xmin)

    #print(strWillWrite)
    f = open('./yolov4Label.txt',"a")
    #写txt
    f.write(strWillWrite)
    #print(strWillWrite)
    if i%100 == 0:
        print(i,'/',len(fileNameList))
    #关闭txt文件
    f.close()
    #print('ok')
print('all finish!')
